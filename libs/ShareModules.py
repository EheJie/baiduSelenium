# coding:utf-8
import os
import logging
from openpyxl import load_workbook




#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
# 函数/过程名称：InsertLog
# 函数/过程的目的：写日志
#-------------------------------------------------------------------------------
#当前文件路径
Curent_dir = os.path.curdir
def InsertLog(level,msg):
    #LogConfFile = './conf/log.conf'
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(filename)-9s [line: %(lineno)3d] " \
                        "%(levelname)-7s %(message)s ", datefmt="%Y/%m/%d %H:%M:%S",\
                        filename=Curent_dir + "/logs/runninglog.log", filemode="a")
    logger = logging.getLogger()
    if level=='INFO' or level=='info':
        logger.info(msg)
    elif level=='ERROR' or level=='error':
        logger.error(msg)
    elif level=='WARNIG' or level=='warnig':
        logger.warning(msg)
    elif level=='DEBUG' or level=='debug':
        logger.debug(msg)
    else:
        logger.critical(msg)

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
# 函数/过程名称：ScreenshotImage
# 函数/过程的目的：截图
#-------------------------------------------------------------------------------
def ScreenshotImage(Driver,FileName):
    try:
        file_path =  Curent_dir + "/result/img/" + FileName
        Driver.get_screenshot_as_file(file_path)
    except BaseException as msg:
        InsertLog('error',msg)

# -------------------------------------------------------------------------------
# ###############################################################################
# -------------------------------------------------------------------------------
# 函数/过程名称：GetSkipScripts
# 函数/过程的目的：获取不需要执行的模块名字
#-------------------------------------------------------------------------------

def GetSkipScripts(FilePath,modelNameColumnIndex,ifRunColumnIndex):
    try:
        m = []
        wb = load_workbook(FilePath)
        ws = wb['ScriptConfig']
        rowcount = ws.max_row
        for i in range(2,rowcount+1):
            cellvalue = ws.cell(row=i,column=ifRunColumnIndex).value
            if cellvalue=='False':
                modulename = ws.cell(row=i,column=modelNameColumnIndex).value
                m.append(modulename)
        wb.close()
        return m
    except BaseException as msg:
        InsertLog('error',msg)

  
# -------------------------------------------------------------------------------
# ###############################################################################
# -------------------------------------------------------------------------------
# 函数/GetSkipTestCases
# 函数/过程的目的：获取不需要执行的用例
#-------------------------------------------------------------------------------
def GetSkipTestCases(FilePath,caseIdColumnIndex,ifRunColumIndex):
    try:
	    #创建一个空列表，用来接收不执行用例的名字
        t = []
		#读取Excel文件
        wb = load_workbook(FilePath)
		#获取Excel所有表格名字
        sheels = wb.sheetnames
        #print sheels
		#第一层for循环，用于遍历所有表格
        for i in sheels:
		    #获取表格名字
            ws = wb[i]
			#获取表格已使用行数
            rowcount = ws.max_row
			#第二层for循环，用于遍历表格每一行数据(从第二行开始)
            for j in range(2,rowcount+1):
			    #获取单元格信息
                cellvalue = ws.cell(row=j,column=ifRunColumIndex).value
				#判断获取到单元格数据的值是否为‘False’
                if cellvalue=='False':
                    testcasename = ws.cell(row=j,column=caseIdColumnIndex).value
                    t.append(testcasename)
        wb.close()
        return t
    except BaseException as msg:
        InsertLog("error",msg)



if __name__ == '__main__':
    pass